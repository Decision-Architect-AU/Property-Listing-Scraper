#!/usr/bin/env python3
"""
build_excel.py  —  Three-Pillar Excel Report Builder
Reads scored_listings.json → SEQ_Listings.xlsx

Usage:
    python build_excel.py
    (must be run from the project directory containing scored_listings.json)
"""

import json
import sys
from datetime import date
from pathlib import Path

# ── Exclusion filters (applied at report time) ────────────────────────────────
CHARACTER_KEYWORDS = [
    "character home", "character property", "character house", "character cottage",
    "character residence", "character building", "character dwelling",
    "unique home", "unique property", "unique character", "truly unique",
    "period home", "period property", "period features", "period charm",
    "art deco", "heritage listed", "heritage home", "heritage property",
    "original character", "original charm", "original features",
    "one of a kind", "timeless charm",
    "built in the 1800", "built in late 1800", "built in the 1900",
]


def is_character_home(listing: dict) -> bool:
    text = " ".join([
        listing.get("description", ""),
        listing.get("listing_description", ""),
    ]).lower()
    return any(kw in text for kw in CHARACTER_KEYWORDS)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────
C_TITLE_BG   = "1A1A2E"   # dark navy
C_TITLE_FG   = "E8E8E8"
C_HEADER_BG  = "16213E"
C_HEADER_FG  = "FFFFFF"
C_GROWTH_BG  = "1B4332"   # deep green
C_GROWTH_FG  = "D8F3DC"
C_DEALS_BG   = "7B2D00"   # deep amber
C_DEALS_FG   = "FFF3E0"
C_CASH_BG    = "0D3B66"   # deep blue
C_CASH_FG    = "E3F2FD"
C_ROW_ALT    = "F8F9FA"
C_ROW_NORM   = "FFFFFF"
C_DELINQUENT = "FFCCCC"
C_ACCENT     = "E8C547"
C_GREEN_HI   = "40916C"
C_DEALS_HI   = "E85D04"
C_CASH_HI    = "0077B6"
C_FAST50     = "7B2FBE"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11, name="Calibri") -> Font:
    return Font(bold=bold, color=color, size=size, name=name)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_thin() -> Border:
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _freeze(ws, cell="B3"):
    ws.freeze_panes = cell


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def _write_all_listings(ws, listings: list[dict], today: str):
    count = len(listings)

    # Title row
    ws.merge_cells("A1:AD1")
    c = ws["A1"]
    c.value = f"All Listings — SEQ Houses Under $2M  |  {count} properties  |  Scraped {today}"
    c.fill  = _fill(C_TITLE_BG)
    c.font  = _font(bold=True, color=C_TITLE_FG, size=13)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # Header row
    headers = ["#", "Suburb", "Address", "Price", "Beds", "Baths", "Park", "Land",
               "Type", "Fast 50",
               "🌱 Growth", "Growth Signals",
               "💰 Deals",  "Deals Signals",
               "💵 Cashflow", "Cashflow Signals",
               "⚠️ Delinquent", "Signals",
               "Listed", "DOM", "Scraped", "URL", "Description",
               # Rent & yield (cols X–AD)
               "Annual Rent $", "Wk Rent $", "Rent Assumptions", "Gross Yield %",
               "Sub Rent/wk", "Sub Median $", "Vacancy %", "SQM Rating"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill      = _fill(C_HEADER_BG)
        c.font      = _font(bold=True, color=C_HEADER_FG)
        c.alignment = _align("center")
        c.border    = _border_thin()
    ws.row_dimensions[2].height = 22

    # Data rows
    for ri, lst in enumerate(listings, 3):
        is_delinquent = lst.get("delinquent", False)
        row_fill = _fill(C_DELINQUENT) if is_delinquent else _fill(C_ROW_ALT if ri % 2 else C_ROW_NORM)

        values = [
            ri - 2,
            lst["suburb"],
            lst["address"],
            lst["price"],
            lst["beds"],
            lst["baths"],
            lst["parking"],
            lst["land"],
            lst["type"],
            "Yes ⭐" if lst["fast_50"] else "No",
            lst["growth_score"],
            lst["growth_signals"],
            lst["deals_score"],
            lst["deals_signals"],
            lst["cashflow_score"],
            lst["cashflow_signals"],
            "⚠️ YES" if is_delinquent else "—",
            lst["delinquent_signal"],
            lst.get("date_listed", ""),
            lst.get("days_on_market", ""),
            lst["scraped_date"],
            lst["url"],
            lst.get("listing_description", ""),
            # Rent & yield
            lst.get("annual_rent") or "",
            lst.get("weekly_rent") or "",
            lst.get("rent_assumptions", ""),
            lst.get("gross_yield_pct") or "",
            lst.get("sub_median_rent_pw") or "",
            lst.get("sub_median_price") or "",
            lst.get("vacancy_pct") or "",
            lst.get("sqm_rating") or "",
        ]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = row_fill
            c.alignment = _align("left", wrap=(ci in (12, 14, 16, 18, 23, 26)))
            c.border    = _border_thin()
            if ci == 11:  # growth score
                c.font = _font(bold=True, color=C_GREEN_HI if (val or 0) >= 5 else "000000")
            elif ci == 13:  # deals score
                c.font = _font(bold=True, color=C_DEALS_HI if (val or 0) >= 3 else "000000")
            elif ci == 15:  # cashflow score
                c.font = _font(bold=True, color=C_CASH_HI if (val or 0) >= 3 else "000000")
            elif ci == 20:  # DOM — colour-code stale listings
                dom_val = val if isinstance(val, int) else 0
                if dom_val >= 180:
                    c.font = _font(bold=True, color="CC0000")
                elif dom_val >= 90:
                    c.font = _font(bold=True, color="E85D04")
                elif dom_val >= 60:
                    c.font = _font(bold=True, color="F4A261")
            elif ci == 27:  # Gross Yield % — colour-code by yield quality
                yld = float(val) if val else 0.0
                if yld >= 7.0:
                    c.font = _font(bold=True, color="1B7A34")   # dark green — excellent
                elif yld >= 5.5:
                    c.font = _font(bold=True, color=C_GREEN_HI)  # green — good
                elif yld >= 4.0:
                    c.font = _font(bold=True, color="E8A000")   # amber — ok
                elif yld > 0:
                    c.font = _font(color="999999")               # grey — low

    # Column widths
    # A# B-Suburb C-Address D-Price E-Beds F-Baths G-Park H-Land I-Type J-Fast50
    # K-Growth L-GrowthSig M-Deals N-DealsSig O-Cash P-CashSig
    # Q-Delinquent R-DelSig S-Listed T-DOM U-Scraped V-URL W-Description
    # X-AnnualRent Y-WkRent Z-RentAssumptions AA-GrossYield%
    # AB-SubRent/wk AC-SubMedian$ AD-Vacancy% AE-SQMRating
    _set_col_widths(ws, {
        "A": 5,  "B": 16, "C": 38, "D": 26, "E": 6,  "F": 6,  "G": 6,
        "H": 10, "I": 12, "J": 8,
        "K": 9,  "L": 40, "M": 8,  "N": 36, "O": 10, "P": 40,
        "Q": 12, "R": 30, "S": 12, "T": 7,  "U": 12, "V": 55, "W": 80,
        "X": 13, "Y": 10, "Z": 36, "AA": 12,
        "AB": 12, "AC": 14, "AD": 10, "AE": 10,
    })
    _freeze(ws, "D3")
    ws.auto_filter.ref = f"A2:AE{len(listings) + 2}"


def _write_pillar_sheet(ws, listings: list[dict], pillar: str, today: str):
    """Generic sheet for Top Growth / Top Deals / Top Cashflow."""
    key_map = {
        "growth":   ("growth_score",   "growth_signals",   "🌱 Growth Score", C_GROWTH_BG, C_GROWTH_FG, C_GREEN_HI),
        "deals":    ("deals_score",    "deals_signals",    "💰 Deals Score",  C_DEALS_BG,  C_DEALS_FG,  C_DEALS_HI),
        "cashflow": ("cashflow_score", "cashflow_signals", "💵 Cashflow Score", C_CASH_BG,  C_CASH_FG,  C_CASH_HI),
    }
    score_key, sig_key, score_label, bg, fg, accent = key_map[pillar]
    emoji = {"growth": "🌱", "deals": "💰", "cashflow": "💵"}[pillar]

    filtered = sorted(
        [l for l in listings if (l.get(score_key) or 0) > 0],
        key=lambda x: x.get(score_key, 0),
        reverse=True,
    )
    count = len(filtered)

    # Title
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f"{emoji} Top {pillar.title()}  |  Sorted by {score_label}  |  {count} scored listings  |  Re-run build_excel.py to refresh"
    c.fill  = _fill(bg)
    c.font  = _font(bold=True, color=fg, size=12)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    # Sub-header
    ws.merge_cells("A2:N2")
    c = ws["A2"]
    c.value = f"  Sorted by {score_label} (highest first)  ·  ⚠️ red rows = delinquent flag  ·  Only listings with score > 0 shown"
    c.fill  = _fill(C_HEADER_BG)
    c.font  = _font(color="AAAAAA", size=10)
    c.alignment = _align("left")
    ws.row_dimensions[2].height = 18

    # Headers
    headers = ["Rank", "Suburb", "Address", "Price", "Land", "Beds", "Type",
               score_label, "Key Signals", "🌱 Growth", "💰 Deals", "💵 Cashflow",
               "⚠️ Delinquent", "Domain URL"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill      = _fill(bg)
        c.font      = _font(bold=True, color=fg)
        c.alignment = _align("center")
        c.border    = _border_thin()
    ws.row_dimensions[3].height = 20

    for ri, lst in enumerate(filtered, 4):
        is_del = lst.get("delinquent", False)
        row_fill = _fill(C_DELINQUENT) if is_del else _fill(C_ROW_ALT if ri % 2 else C_ROW_NORM)
        values = [
            ri - 3,
            lst["suburb"],
            lst["address"],
            lst["price"],
            lst["land"],
            lst["beds"],
            lst["type"],
            lst.get(score_key, 0),
            lst.get(sig_key, "—"),
            lst["growth_score"],
            lst["deals_score"],
            lst["cashflow_score"],
            "⚠️ YES" if is_del else "—",
            lst["url"],
        ]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = row_fill
            c.alignment = _align("left", wrap=(ci == 9))
            c.border    = _border_thin()
            if ci == 8:
                c.font = _font(bold=True, color=accent, size=12)

    _set_col_widths(ws, {
        "A": 6, "B": 16, "C": 40, "D": 26, "E": 10, "F": 6, "G": 14,
        "H": 13, "I": 48, "J": 9, "K": 9, "L": 10, "M": 12, "N": 55,
    })
    _freeze(ws, "C4")


def _write_zoning_sheet(ws):
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Zoning Reference — SEQ Suburbs  |  Logan City Council, Ipswich City Council & Moreton Bay"
    c.fill  = _fill(C_TITLE_BG)
    c.font  = _font(bold=True, color=C_TITLE_FG, size=12)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    headers = ["Suburb", "LGA / Council", "Zoning Type", "Zone Notes", "Growth Potential", "Fast 50"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill      = _fill(C_HEADER_BG)
        c.font      = _font(bold=True, color=C_HEADER_FG)
        c.alignment = _align("center")
        c.border    = _border_thin()

    from regions import ZONING, REZONE_POTENTIAL, FAST_50
    for ri, (suburb, info) in enumerate(sorted(ZONING.items()), 3):
        rp = REZONE_POTENTIAL.get(suburb, "UNKNOWN")
        f50 = "Yes ⭐" if suburb in FAST_50 else "No"
        values = [suburb, info["council"], info["zone"], info["notes"], rp, f50]
        row_fill = _fill(C_ROW_ALT if ri % 2 else C_ROW_NORM)
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = row_fill
            c.alignment = _align("left", wrap=(ci in (4, 5)))
            c.border    = _border_thin()

    _set_col_widths(ws, {"A": 18, "B": 24, "C": 28, "D": 52, "E": 40, "F": 8})


def _write_dashboard(ws, listings: list[dict], today: str):
    total = len(listings)
    growth_scored  = sum(1 for l in listings if l.get("growth_score",   0) > 0)
    deals_scored   = sum(1 for l in listings if l.get("deals_score",    0) > 0)
    cash_scored    = sum(1 for l in listings if l.get("cashflow_score", 0) > 0)
    delinquent_ct  = sum(1 for l in listings if l.get("delinquent", False))
    fast50_ct      = sum(1 for l in listings if l.get("fast_50", False))
    top_growth     = sorted(listings, key=lambda x: x.get("growth_score", 0),   reverse=True)[:5]
    top_deals      = sorted(listings, key=lambda x: x.get("deals_score", 0),    reverse=True)[:5]
    top_cash       = sorted(listings, key=lambda x: x.get("cashflow_score", 0), reverse=True)[:5]

    ws.merge_cells("A1:S1")
    c = ws["A1"]
    c.value = f"SEQ Property Investment Pipeline  |  Three-Pillar Strategy  |  Scraped {today}  |  {total} listings"
    c.fill  = _fill(C_TITLE_BG)
    c.font  = _font(bold=True, color=C_ACCENT, size=15)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 34

    # Strategy map header
    ws.merge_cells("A3:S3")
    c = ws["A3"]
    c.value = "  PILLAR STRATEGY MAP"
    c.fill  = _fill(C_HEADER_BG)
    c.font  = _font(bold=True, color=C_ACCENT, size=12)
    c.alignment = _align("left")

    # Three pillar summary boxes  (rows 5-9)
    pillars = [
        ("A", "🌱 GROWTH",   f"{growth_scored} listings scored",  C_GROWTH_BG, C_GROWTH_FG),
        ("G", "💰 DEALS",    f"{deals_scored} listings scored",    C_DEALS_BG,  C_DEALS_FG),
        ("M", "💵 CASHFLOW", f"{cash_scored} listings scored",     C_CASH_BG,   C_CASH_FG),
    ]
    for col, label, sub, bg, fg in pillars:
        end = chr(ord(col) + 4)
        ws.merge_cells(f"{col}5:{end}5")
        c = ws[f"{col}5"]
        c.value = label;  c.fill = _fill(bg); c.font = _font(bold=True, color=fg, size=13); c.alignment = _align("center")
        ws.merge_cells(f"{col}6:{end}6")
        c = ws[f"{col}6"]
        c.value = sub;    c.fill = _fill(bg); c.font = _font(color=fg, size=10);           c.alignment = _align("center")

    # Stats row
    ws.merge_cells("A8:F8")
    ws["A8"].value = f"📊 Total listings: {total}    ⚠️ Delinquent: {delinquent_ct}    ⭐ Fast 50: {fast50_ct}"
    ws["A8"].font  = _font(bold=True, color=C_TITLE_BG, size=11)
    ws["A8"].fill  = _fill("F0F0F0")

    # Top 5 tables per pillar
    def _top5_table(start_row, items, score_key, label, bg, fg, acc):
        ws.merge_cells(f"A{start_row}:F{start_row}")
        c = ws[f"A{start_row}"]
        c.value = f"  🏆 Top 5 {label}"; c.fill = _fill(bg); c.font = _font(bold=True, color=fg, size=11); c.alignment = _align("left")

        ws.merge_cells(f"G{start_row}:L{start_row}")
        c = ws[f"G{start_row}"]
        c.value = f"  🏆 Top 5 Deals"; c.fill = _fill(C_DEALS_BG); c.font = _font(bold=True, color=C_DEALS_FG, size=11); c.alignment = _align("left")

        ws.merge_cells(f"M{start_row}:S{start_row}")
        c = ws[f"M{start_row}"]
        c.value = f"  🏆 Top 5 Cashflow"; c.fill = _fill(C_CASH_BG); c.font = _font(bold=True, color=C_CASH_FG, size=11); c.alignment = _align("left")

        for idx, (tg, td, tc) in enumerate(zip(top_growth, top_deals, top_cash)):
            r = start_row + 1 + idx
            rf = _fill(C_ROW_ALT if idx % 2 else C_ROW_NORM)

            def _write_row(col_start, lst, sk, ac):
                cols = list(range(col_start, col_start + 6))
                vals = [idx + 1, lst["suburb"], lst["address"][:30], lst["price"][:20], lst.get(sk, 0), ""]
                for ci, val in zip(cols, vals):
                    cell = ws.cell(row=r, column=ci, value=val)
                    cell.fill = rf
                    cell.border = _border_thin()
                    if ci == cols[4]:
                        cell.font = _font(bold=True, color=ac)

            _write_row(1, tg, "growth_score",   C_GREEN_HI)
            _write_row(7, td, "deals_score",    C_DEALS_HI)
            _write_row(13, tc, "cashflow_score", C_CASH_HI)

    _top5_table(10, top_growth, "growth_score", "Growth", C_GROWTH_BG, C_GROWTH_FG, C_GREEN_HI)

    _set_col_widths(ws, {
        "A": 5, "B": 16, "C": 30, "D": 22, "E": 8, "F": 4,
        "G": 5, "H": 16, "I": 30, "J": 22, "K": 8, "L": 4,
        "M": 5, "N": 16, "O": 30, "P": 22, "Q": 8, "R": 4, "S": 4,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main(scored_path: Path = None, out_path: Path = None):
    """
    Build Excel report from scored_listings.json.

    Args:
        scored_path: Path to scored_listings.json (default: project dir)
        out_path:    Output .xlsx path (default: SEQ_Listings.xlsx in project dir)
    """
    project_dir = Path(__file__).parent
    if scored_path is None:
        scored_path = project_dir / "scored_listings.json"
    if out_path is None:
        out_path = project_dir / "SEQ_Listings.xlsx"

    scored_path = Path(scored_path)
    out_path    = Path(out_path)

    if not scored_path.exists():
        print(f"ERROR: {scored_path} not found — run score_listings.py first", file=sys.stderr)
        sys.exit(1)

    all_listings = json.loads(scored_path.read_text(encoding="utf-8"))
    # Exclude sold/delisted/inactive and character homes from the report
    listings = [l for l in all_listings
                if l.get("active", True) and not is_character_home(l)]
    today    = str(date.today())
    sold_ct  = len(all_listings) - len(listings)
    print(f"Building Excel for {len(listings)} active listings ({sold_ct} sold/inactive excluded) -> {out_path}")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Sheet order
    ws_dash  = wb.create_sheet("📊 Dashboard")
    ws_all   = wb.create_sheet("📋 All Listings")
    ws_grow  = wb.create_sheet("🌱 Top Growth")
    ws_deals = wb.create_sheet("💰 Top Deals")
    ws_cash  = wb.create_sheet("💵 Top Cashflow")
    ws_zone  = wb.create_sheet("🗺 Zoning Reference")

    _write_dashboard(ws_dash, listings, today)
    _write_all_listings(ws_all, listings, today)
    _write_pillar_sheet(ws_grow,  listings, "growth",   today)
    _write_pillar_sheet(ws_deals, listings, "deals",    today)
    _write_pillar_sheet(ws_cash,  listings, "cashflow", today)
    _write_zoning_sheet(ws_zone)

    wb.save(str(out_path))
    print(f"Saved listings: {out_path}")


if __name__ == "__main__":
    main()
